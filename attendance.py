import face_recognition
import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import calendar
import os
import eel
import base64
import numpy as np

eel.init('web')

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
DIRECTORY_PATH = PROJECT_ROOT
IMAGE_DIRECTORY_PATH = os.path.join(PROJECT_ROOT, 'assets')
os.makedirs(IMAGE_DIRECTORY_PATH, exist_ok=True)


#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 1 (SAVING THE IMAGE DATA FROM JS -> PYTHON -> LOCAL STORAGE)
#-------------------------------------------------------------------------------------------------------------------------------------#

# Define Python function called in JS and expose it to eel, this method will save the screenshot image data with name image.jpg

@eel.expose
def get_image_data(data):
    header, encoded = data.split(',', 1)
    with open(os.path.join(DIRECTORY_PATH, 'image.jpg'), 'wb') as f:
        f.write(base64.b64decode(encoded))
    return "Python says thanks"




#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 2 (CREATING EXCEL SHEET AND CALCULATING DEPENDENCIES)
#-------------------------------------------------------------------------------------------------------------------------------------#

# Load everything i.e workbook, video capture (videocam)

# Load present date and time
now = datetime.datetime.now()
today = now.day
month = now.month
year = now.year
month_name = now.strftime('%B')
days_in_month = calendar.monthrange(year, month)[1]

attendance_rows = {}

# Create a worksheet only when the file does not exist else update the existing excel workbook
workbook_path = os.path.join(DIRECTORY_PATH, str(month_name) + '.xlsx')
if not os.path.isfile(workbook_path):
    book = Workbook()
    sheet = book.active
else:
    book = openpyxl.load_workbook(workbook_path)
    sheet = book.active


def _style_attendance_sheet():
    global sheet

    title = "SMART ATTENDANCE SYSTEM"
    subtitle = f"Attendance Register — {month_name} {year}"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1D4ED8")
    body_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    legacy_entries = []
    if sheet['A1'].value != title:
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value not in (None, ""):
                    legacy_entries.append((row_idx, col_idx, cell.value))

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col_idx).value = None

    sheet.title = f"Attendance_{month_name}_{year}"
    sheet.cell(row=1, column=1).value = title
    sheet.cell(row=2, column=1).value = subtitle

    sheet.cell(row=4, column=1).value = "Roll No."
    for day in range(1, days_in_month + 1):
        sheet.cell(row=4, column=day + 1).value = f"{day:02d}"

    for row_idx in range(4, 5):
        for col_idx in range(1, days_in_month + 2):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

    for row_idx in range(5, 1000):
        sheet.row_dimensions[row_idx].height = 22

    for col_idx in range(1, days_in_month + 2):
        if col_idx == 1:
            sheet.column_dimensions["A"].width = 16
        else:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 5.5

    sheet.freeze_panes = "B5"

    if legacy_entries:
        row_lookup = {}
        for row_idx in range(5, sheet.max_row + 1):
            roll_value = sheet.cell(row=row_idx, column=1).value
            if roll_value not in (None, ""):
                if isinstance(roll_value, int) or str(roll_value).isdigit():
                    row_lookup[int(roll_value)] = row_idx

        for legacy_row, legacy_col, legacy_value in legacy_entries:
            if legacy_row in (1, 2, 3, 4):
                continue
            if legacy_col < 1 or legacy_col > days_in_month:
                continue
            if not isinstance(legacy_row, int) or isinstance(legacy_row, bool):
                continue

            roll_number = legacy_row
            day_number = legacy_col
            if roll_number <= 0 or day_number <= 0:
                continue

            if roll_number not in row_lookup:
                next_row = len(row_lookup) + 5
                row_lookup[roll_number] = next_row
                sheet.cell(row=next_row, column=1).value = roll_number
                sheet.cell(row=next_row, column=1).font = Font(bold=True)
                sheet.cell(row=next_row, column=1).alignment = center_alignment
                sheet.cell(row=next_row, column=1).border = border
            target_row = row_lookup[roll_number]
            target_col = day_number + 1
            sheet.cell(row=target_row, column=target_col).value = legacy_value
            sheet.cell(row=target_row, column=target_col).alignment = center_alignment
            sheet.cell(row=target_row, column=target_col).border = border
            if isinstance(legacy_value, str) and legacy_value.lower() in ("present", "p"):
                sheet.cell(row=target_row, column=target_col).fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
                sheet.cell(row=target_row, column=target_col).font = Font(bold=True, color="047857")

    for row_idx in range(5, 1000):
        roll_value = sheet.cell(row=row_idx, column=1).value
        if roll_value in (None, ""):
            break
        attendance_rows[int(roll_value)] = row_idx
        for col_idx in range(2, days_in_month + 2):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value in (None, ""):
                cell.border = border
                cell.alignment = center_alignment
                continue
            cell.border = border
            cell.alignment = center_alignment
            if cell.value == "Present":
                cell.fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
                cell.font = Font(bold=True, color="047857")
            elif cell.value == "Absent":
                cell.fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
                cell.font = Font(bold=True, color="B91C1C")
            else:
                cell.fill = body_fill


def _ensure_attendance_row(roll_number):
    global sheet

    if roll_number in attendance_rows:
        return attendance_rows[roll_number]

    next_row = len(attendance_rows) + 5
    attendance_rows[roll_number] = next_row
    sheet.cell(row=next_row, column=1).value = roll_number
    sheet.cell(row=next_row, column=1).font = Font(bold=True)
    sheet.cell(row=next_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=next_row, column=1).border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    return next_row


def _mark_attendance(roll_number, status):
    global sheet

    row_idx = _ensure_attendance_row(roll_number)
    target_col = today + 1
    cell = sheet.cell(row=row_idx, column=target_col)
    cell.value = status
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    if status == "Present":
        cell.fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
        cell.font = Font(bold=True, color="047857")
    elif status == "Absent":
        cell.fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
        cell.font = Font(bold=True, color="B91C1C")
    else:
        cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        cell.font = Font(bold=False)


_style_attendance_sheet()
book.save(workbook_path)




#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 3 (TRAINING AND LOADING KNOWN FACE ENCODINGS)
#-------------------------------------------------------------------------------------------------------------------------------------#

@eel.expose
def save_student_data(roll_number):
    source_image = os.path.join(DIRECTORY_PATH, 'image.jpg')
    target_image = os.path.join(IMAGE_DIRECTORY_PATH, str(roll_number) + '.jpg')

    if not os.path.isfile(source_image):
        return 'No captured image found. Click Take Picture before Add Student.'

    os.replace(source_image, target_image)    # This will change the image.jpg to the roll number inserted in field.jpg
    return 'Student image saved.'

        

# Create arrays of known face encodings and their names
known_face_encodings = []
known_face_names = []

print('Following images have been trained : ')
for filename in os.listdir(IMAGE_DIRECTORY_PATH):
    print(str(filename))
    if filename.endswith(".jpg"):
        image = face_recognition.load_image_file(os.path.join(IMAGE_DIRECTORY_PATH, str(filename)))                 # Load Image
        image = np.ascontiguousarray(image)
        image_face_encoding = face_recognition.face_encodings(image)[0]         # Find the face encoding
        known_face_encodings.append(image_face_encoding)                        # Append the face encoding to the known faces
        known_face_names.append(filename)                                       # Append the name to the known face names
    



#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 4 (STARTING THE WEB CAM STREAM FOR TAKING ATTENDANCE, PRESS CTRL + C TO STOP )
#-------------------------------------------------------------------------------------------------------------------------------------#

# This method will be used starting the video stream for taking the attendance of all the already known faces

@eel.expose
def take_attendance():
    video_capture = cv2.VideoCapture(0)

    if not video_capture.isOpened():
        print("Could not open the webcam for attendance.")
        eel.show_notification(
            "error",
            "Attendance Could Not Be Recorded",
            "Please check the system and try again.",
        )
        return

    # Initialize some variables
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True
    recorded_rolls = set()
        
    
    while True:
        
        # Grab a single frame of video
        ret, frame = video_capture.read()
        
        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        
        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = np.ascontiguousarray(small_frame[:, :, ::-1])
        
        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame)
            try:
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations, num_jitters=0)
            except TypeError as error:
                print("Face encoding failed for this frame:", error)
                face_encodings = []
        
        face_names = []
        name = ""
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
        
            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                name = name[:name.find('.')]
                # Assign attendance
                if int(name) in range(1,61):
                    roll_number = int(name)
                    if roll_number not in recorded_rolls:
                        _mark_attendance(roll_number, "Present")
                        recorded_rolls.add(roll_number)
                        date_label = datetime.datetime.now().strftime('%d %B %Y')
                        eel.show_notification(
                            "success",
                            "Attendance Recorded",
                            f"Roll No: {roll_number}\nDate: {date_label}\nStatus: Present",
                        )
                else:
                    pass
        
        face_names.append(name)
        
        process_this_frame = not process_this_frame
        
        top, right, left, bottom = 1,1,1,1      # Default values since the variables not defined were giving error
        # Display the results
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
        
        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        
            # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
        
        # Display the resulting image
        cv2.imshow('Video', frame)
            
        # Save Woorksheet as present month
        book.save(os.path.join(DIRECTORY_PATH, str(month_name) + '.xlsx'))
        
        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        
    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()




#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 5 (EEL STARTING THE SERVER)
#-------------------------------------------------------------------------------------------------------------------------------------#

# To start the web server of the eel 

eel.start('index.html', mode='chrome-app', port=8080, cmdline_args=['--start-fullscreen', '--browser-startup-dialog'])